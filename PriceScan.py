#! python 3

from requests import get
from os import chdir
from bs4 import BeautifulSoup
import re
import logging
import openpyxl
from openpyxl.styles import PatternFill

# ## FLOW ##
#   Initialization - establish global variables / values ...
#  STEP 1. Main Loop.  Open the File / DB with the list of items to price check and loop thru each entry
#  STEP 2. Retrieve the WebPage to memory
#  STEP 3. Extract the webpage section based on the CSS Selector Path
#  STEP 4. Parse the "retrieved webpage section" to get the data
#  STEP 5. Update tine File / DB with the latest price

def initRoutine():
  from ruamel.yaml import YAML
  global CSS_Selector, ShortNameLength, LogFileName, LogDisableLevel, WorkBookFileName, WorksheetTabName, Item_Column, WebSiteURL_Column, CSS_SelectorColumn, OldPrice_Column, NewPrice_Column, DataStart_Row

  try:
    ymlfile = open("config.yml", 'r')  # open the YAML file and assign it to variable 
    yaml=YAML(typ='safe')   # default, if not specfied, is 'rt' (round-trip)
    cfg = yaml.load(ymlfile)

    ymlGeneral = cfg['General']
    CSS_Selector = ymlGeneral['CSS_SelectorDefault']
    ShortNameLength = ymlGeneral['ShortBookNameLength']

    ymlLogging = cfg['Logging']
    LogFileName = ymlLogging['LogFileName']
    LogDisableLevel = ymlLogging['Disable_Level']

    # Use logging module's basicConfig, disable and info methods
    logging.basicConfig(filename=LogFileName, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
    if LogDisableLevel == 'DEBUG':
      logging.disable(logging.DEBUG)  # when you want to Turn off low level logging
    elif LogDisableLevel == 'INFO':
      logging.disable(logging.INFO)  # when you want to Turn off low level logging
    elif LogDisableLevel == 'WARNING':
      logging.disable(logging.WARNING)  # when you want to Turn off low level logging
    elif LogDisableLevel == 'ERROR':
      logging.disable(logging.ERROR)  # when you want to Turn off low level logging
    elif LogDisableLevel == 'CRITICAL':
      logging.disable(logging.CRITICAL)  # when you want to Turn off low level logging
    logging.critical('+ + + + + Program Started - Initialization Routine + + + + +')

    ymlExcelType = cfg['ExcelType']
    WorkBookFileName = ymlExcelType['WorkBookFileName']
    WorksheetTabName = ymlExcelType['WorksheetTabName']
    Item_Column = ymlExcelType['Item_Column']
    WebSiteURL_Column = ymlExcelType['WebSiteURL_Column']
    CSS_SelectorColumn = ymlExcelType['CSS_SelectorColumn']
    OldPrice_Column = ymlExcelType['OldPrice_Column']
    NewPrice_Column = ymlExcelType['NewPrice_Column']
    DataStart_Row = ymlExcelType['DataStart_Row']

    sConfigData = '''* * * Configuration data loaded from config.yml:
    \tGeneral Setting - CSS_SelectorDefault: %s 
    \tGeneral Setting - ShortBookNameLength: %s 
    \tLogging - LogFileName: %s 
    \tLogging - Disable_Level: %s 
    \tExcelType - WorkBookFileName: %s 
    \tExcelType - WorksheetTabName: %s 
    \tExcelType - Item_Column: %s 
    \tExcelType - WebSiteURL_Column: %s 
    \tExcelType - CSS_SelectorColumn: %s 
    \tExcelType - OldPrice_Column: %s 
    \tExcelType - NewPrice_Column: %s 
    \tExcelType - DataStart_Row: %s 
    * * * End of Config data 
    '''%(CSS_Selector, ShortNameLength, LogFileName, LogDisableLevel, WorkBookFileName, WorksheetTabName, Item_Column, WebSiteURL_Column, CSS_SelectorColumn, OldPrice_Column, NewPrice_Column, DataStart_Row)

    logging.critical(sConfigData)
    initSuccess = True

  except FileNotFoundError:
    print("Config file not found")
    LogFileName = 'myLogFile'
    logging.basicConfig(filename=LogFileName, level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
    sError = '+ + + + + Program Aborted - Initialization Routine + + + + + \n'
    sError = sError + ' # # # # #  Configuration file not found # # # # #'
    logging.critical(sError)
    initSuccess = False

  return(initSuccess)


def getWebPage(webPageURL, CSS_Sel1, bookname):
  # Use the Requests module get method to retrieve a web page into memory.
  global iWebPage, iWebCSS
  elems = []
  myResponse = get(webPageURL)

  try:
    myResponse.raise_for_status()
    blFoundPage = True
  except Exception as exc:
    logging.error('WebPage %s had this error: %s' %(webPageURL, exc))
    blFoundPage = False

  if myResponse.status_code == 200:
    logging.debug('FOUND website %s' %webPageURL)

  if blFoundPage:
    myParse = BeautifulSoup(myResponse.text, 'html.parser')   # Use the bs4 module BeautifulSoup method to first create the Search object, then to search it with the search string
    elems = myParse.select(CSS_Sel1)

    if elems != []:
      logging.debug(str(len(elems)) + ' Strings found on page.')
    else:
      logging.warning('There were no strings found when searching for %s in %s' %(CSS_Sel1, bookname))
      iWebCSS += 1

          # You failed to find the CSS Selector.  If config file specifies, here is where you save the downloaded file for debugging
          # TODO:  Figure out how you'll name the page and add config item for the Chunk size
          # savedFile = open(downLoadFile,'wb')
          # for myChunk in myResponse.iter_content(100000):
          #   savedFile.write(myChunk)
          # savedFile.close()

  else:
    iWebPage += 1

  # TODO: Pass back success / fail status on each CSS_Selector string
  return(elems)


def parseOutPrice(argLstUnformatLines):
  myRegObj = re.compile(r'\$(\d)+.\d\d')
  myMatch = myRegObj.search(argLstUnformatLines[0].text)
  myPrices = myMatch.group()
  return (myPrices)

# TODO:  Figure out how to determine where the downloaded web pages will get stored.  Make a subdir for this run.  Dir name s/b YYMMDD ...
# Use os Module chdir method to set the working directory
cacheDir = 'K:\\@Tech\\Programming\\Python\\PriceScan'
chdir(cacheDir)

# Initialize Global Variables
iWebPage = 0
iWebCSS = 0
iParse = 0

CSS_Selector = None
LogFileName = None
LogDisableLevel = None

ShortNameLength = None
WorkBookFileName = None
WorksheetTabName = None
Item_Column = None
WebSiteURL_Column = None
CSS_SelectorColumn = None
OldPrice_Column = None
NewPrice_Column = None
DataStart_Row = None

if initRoutine():
  #try:
  wb = openpyxl.load_workbook(WorkBookFileName)
  sheet = wb[WorksheetTabName]
  logging.debug('Input xls file opened.  There were ' + str(sheet.max_column) + ' columns and ' + str(sheet.max_row) + ' rows.')

  x = int(DataStart_Row)
  while x < sheet.max_row + 1:
    parmWebPageURL = sheet.cell(row = x, column = int(WebSiteURL_Column)).value   #  get from input file
    parmCSS_Sel1 = sheet.cell(row = x, column = int(CSS_SelectorColumn)).value     #  get from input file.  Use default from config if null.
    if not parmCSS_Sel1:                                     #  see if this item in file's CSS Selector value is null / empty use default from config file
      parmCSS_Sel1 = CSS_Selector
    wrkBookName = sheet.cell(row = x, column = int(Item_Column)).value     #  get book name from file or increment a counter.
    if len(wrkBookName) == int(ShortNameLength):
      parmBookName = wrkBookName
    elif len(wrkBookName) > int(ShortNameLength):
      parmBookName = wrkBookName[0:int(ShortNameLength)-1]
    else:
      parmBookName = wrkBookName.ljust(int(ShortNameLength), '*')

    lstUnformattedPriceLines = getWebPage(parmWebPageURL, parmCSS_Sel1, parmBookName)   # Pass web page and CSS search strings and get a list of the unformatted strings from webpage
    if lstUnformattedPriceLines != []:
      logging.debug(str(len(lstUnformattedPriceLines)) + ' Strings found on page.')
      strBookPrice = parseOutPrice(lstUnformattedPriceLines)
      if strBookPrice == '':
        iParse += 1
      logging.debug('*****  Price for %s is: %s' %(parmBookName, strBookPrice))
      sheet.cell(row = x, column = int(NewPrice_Column)).value = strBookPrice

      # TODO: compare the parsed price to previous price on input file and do something ...
      # if True:
      #   fill = PatternFill(bgColor='92D050', fill_type = 'solid')   # Green
      # else:
      #   fill = PatternFill(bgColor='FF0000', fill_type = 'solid')   # Red
      # sheet['B2'].fill = fill
    # else:    # Logic to execute when you didn't find web site ore CSS_Selector for this item
    x = x + 1

  # Wrap up processing.  Determine / log job stats.  Save and close Excel I/O file.
  sErrRate = '{:.2%}'.format(((iWebCSS + iWebPage + iParse)/sheet.max_row))
  sJobResults = '''* * * Job processing reults:
  \tProcessed %s rows. 
  \tThere were %s unsuccessful lookups:
  \t\tWebPage Retrieval issues: %s
  \t\tCSS Lookup issues: %s
  \t\tParsing issues: %s
  \t\tFor a %s failure rate. 
  * * * End of Job Results 
  '''%(str(sheet.max_row), str(iWebCSS + iWebPage + iParse), str(iWebPage), str(iWebCSS), str(iParse), sErrRate)
  logging.critical(sJobResults)
  wb.save(WorkBookFileName)
  wb.close()

logging.critical('+ + + + + Program Ended + + + + +')